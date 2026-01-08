import streamlit as st
import pandas as pd
import io
import os

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# =========================
# 업로드 폴더 설정
# =========================
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

title_fill = PatternFill("solid", fgColor="DDDDDD")
value_fill = PatternFill("solid", fgColor="FFFFFF")

# -------------------------
# Streamlit 기본 설정
# -------------------------
st.set_page_config(
    page_title="픽업/샌딩 시간대 월별 누적 집계",
    layout="wide"
)

st.title("🚐 픽업 / 샌딩 시간대 월별 누적 집계 ")

# -------------------------
# 시간 파싱 함수
# -------------------------
def parse_hour(value):
    if pd.isna(value):
        return None
    try:
        value = str(value).strip()
        if value == "":
            return None
        if ":" in value:
            return int(value.split(":")[0])
        return None
    except:
        return None

# =========================
# 업로드된 파일 목록 + 삭제
# =========================
st.subheader("📁 업로드된 파일 목록")

stored_files = sorted([
    f for f in os.listdir(UPLOAD_DIR)
    if f.lower().endswith(".xlsx")
])

if stored_files:
    for file_name in stored_files:
        col1, col2, col3 = st.columns([6, 2, 2])

        # 파일명
        col1.write(f"📄 {file_name}")

        # 다운로드 버튼
        file_path = os.path.join(UPLOAD_DIR, file_name)
        with open(file_path, "rb") as f:
            col2.download_button(
                label="⬇ 다운로드",
                data=f,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{file_name}"
            )

        # 삭제 버튼
        if col3.button("❌ 삭제", key=f"del_{file_name}"):
            os.remove(file_path)
            st.rerun()
else:
    st.info("업로드된 파일이 없습니다.")


# -------------------------
# 파일 업로드
# -------------------------
uploaded_files = st.file_uploader(
    "📂 엑셀 파일 업로드 (연도별 가능, 저장됨)",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    for file in uploaded_files:
        save_path = os.path.join(UPLOAD_DIR, file.name)
        with open(save_path, "wb") as f:
            f.write(file.getbuffer())
    st.success("파일이 저장되었습니다.")
    st.rerun()

# =========================
# 저장된 파일로 통계 처리
# =========================
if stored_files:
    dfs = []

    for fname in stored_files:
        path = os.path.join(UPLOAD_DIR, fname)
        df = pd.read_excel(path)
        df["__source_file"] = fname
        dfs.append(df)

    df = pd.concat(dfs, ignore_index=True)

    # 필수 컬럼 체크
    required_cols = ["출발일", "출발시간", "서비스"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"필수 컬럼이 없습니다: {missing}")
        st.stop()

    # 날짜 / 시간 처리
    df["출발일"] = pd.to_datetime(df["출발일"], errors="coerce")
    df["연월"] = df["출발일"].dt.strftime("%Y-%m")
    df["시간"] = df["출발시간"].apply(parse_hour)

    # 서비스 구분
    df["서비스명"] = df["서비스"].map({"P": "픽업", "S": "샌딩"})
    df = df.dropna(subset=["연월", "시간", "서비스명"])

    # 시간대 라벨 (✅ 00:00 형태로 변경)
    hour_labels = {h: f"{str(h).zfill(2)}:00" for h in range(24)}

    # -------------------------
    # 피벗 테이블 생성 함수
    # -------------------------
    def make_pivot(service_name):
        temp = df[df["서비스명"] == service_name]

        pivot = (
            temp.pivot_table(
                index="연월",
                columns="시간",
                values="서비스명",
                aggfunc="count",
                fill_value=0
            )
            .rename(columns=hour_labels)
        )

        # 모든 시간대 컬럼 보장
        for col in hour_labels.values():
            if col not in pivot.columns:
                pivot[col] = 0

        pivot = pivot[list(hour_labels.values())]
        pivot["총 건수"] = pivot.sum(axis=1)
        pivot = pivot.sort_index(ascending=False)

        # 총합계 행
        total_row = pivot.sum().to_frame().T
        total_row.index = ["총합계"]
        pivot = pd.concat([pivot, total_row])

        return pivot

    pickup_df = make_pivot("픽업")
    sending_df = make_pivot("샌딩")

    # -------------------------
    # 화면 표시
    # -------------------------
    st.subheader("📊 픽업 (누적)")
    st.dataframe(pickup_df, use_container_width=True)

    st.subheader("📊 샌딩 (누적)")
    st.dataframe(sending_df, use_container_width=True)

    # -------------------------
    # 엑셀 다운로드 + 대시보드
    # -------------------------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pickup_df.to_excel(writer, sheet_name="픽업")
        sending_df.to_excel(writer, sheet_name="샌딩")

        wb = writer.book
        ws_pickup = wb["픽업"]
        ws_sending = wb["샌딩"]

        bold = Font(bold=True)

        # 총합계 행 굵게
        for ws in [ws_pickup, ws_sending]:
            last_row = ws.max_row
            for col in range(1, ws.max_column + 1):
                ws.cell(row=last_row, column=col).font = bold

        # 최대 시간대 계산
        pickup_peak = pickup_df.loc["총합계"].drop("총 건수").idxmax()
        sending_peak = sending_df.loc["총합계"].drop("총 건수").idxmax()

        # -------------------------
        # 대시보드 시트
        # -------------------------
        ws_dash = wb.create_sheet("대시보드")

        kpi_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 배경 (연한 회색) - 대시보드 전체 영역
        bg_fill = PatternFill("solid", fgColor="EEEEEE")
        for row in range(1, 81):
            for col in range(1, 41):
                ws_dash.cell(row=row, column=col).fill = bg_fill

        # KPI 함수
        def kpi(title_cell, value_cell, title, value):
            ws_dash.merge_cells(title_cell)
            ws_dash.merge_cells(value_cell)

            title_anchor = title_cell.split(":")[0]
            value_anchor = value_cell.split(":")[0]

            ws_dash[title_anchor] = title
            ws_dash[value_anchor] = value

            ws_dash[title_anchor].font = Font(bold=True)
            ws_dash[value_anchor].font = Font(bold=True, size=16)

            ws_dash[title_anchor].alignment = Alignment(horizontal="center", vertical="center")
            ws_dash[value_anchor].alignment = Alignment(horizontal="center", vertical="center")

            for row in ws_dash[title_cell]:
                for cell in row:
                    cell.fill = title_fill
                    cell.border = kpi_border
            for row in ws_dash[value_cell]:
                for cell in row:
                    cell.fill = value_fill
                    cell.border = kpi_border

        # KPI 배치
        kpi("B2:D2", "B3:D4", "총 픽업 건수", pickup_df.loc["총합계", "총 건수"])
        kpi("F2:H2", "F3:H4", "총 샌딩 건수", sending_df.loc["총합계", "총 건수"])
        kpi("J2:L2", "J3:L4", "픽업 최대 시간대", pickup_peak)
        kpi("N2:P2", "N3:P4", "샌딩 최대 시간대", sending_peak)

        ws_dash.merge_cells("B6:P6")
        ws_dash["B6"] = "픽업 시간별 건수"
        ws_dash["B6"].font = Font(bold=True, size=22)
        ws_dash["B6"].alignment = Alignment(horizontal="center", vertical="center")

        # 배경색 (연회색보다 살짝 진하게)
        title_fill = PatternFill("solid", fgColor="DDDDDD")

        # 테두리 스타일
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 병합된 영역 전체에 배경 + 테두리 적용
        for row in ws_dash["B6:P6"]:
            for cell in row:
                cell.fill = title_fill
                cell.border = title_border


        # -------------------------
        # 픽업 차트
        # -------------------------
        pickup_chart = BarChart()
        pickup_chart.title = None
        pickup_chart.legend = None
        pickup_chart.height = 12
        pickup_chart.width = 28
        pickup_chart.dataLabels = DataLabelList()
        pickup_chart.dataLabels.showVal = True      # 값만 표시
        pickup_chart.dataLabels.showCatName = False # 시간 안씀 (축에만)
        pickup_chart.dataLabels.showSerName = False # ❌ 계열1 제거
        pickup_chart.dataLabels.dLblPos = "outEnd"

        pickup_chart.add_data(
            Reference(
                ws_pickup,
                min_col=2,
                min_row=ws_pickup.max_row,          # 총합계 행
                max_col=ws_pickup.max_column - 1,   # '총 건수' 제외
                max_row=ws_pickup.max_row
            ),
            titles_from_data=False,
            from_rows=True  # ✅ 핵심
        )

        pickup_chart.set_categories(
            Reference(
                ws_pickup,
                min_col=2,
                min_row=1,                          # 시간 헤더
                max_col=ws_pickup.max_column - 1,
                max_row=1
            )
        )

        # ✅ 막대 위 값 표시
        
        pickup_chart.x_axis.tickLblPos = "nextTo"
        pickup_chart.x_axis.majorTickMark = "out"
        pickup_chart.x_axis.minorTickMark = "none"
        pickup_chart.x_axis.delete = False

        ws_dash.add_chart(pickup_chart, "B8")

        # -------------------------
        # 샌딩 차트
        # -------------------------
        ws_dash.merge_cells("B30:P30")
        ws_dash["B30"] = "센딩 시간별 건수"
        ws_dash["B30"].font = Font(bold=True, size=22)
        ws_dash["B30"].alignment = Alignment(horizontal="center", vertical="center")

        # 배경색 (연회색보다 살짝 진하게)
        title_fill = PatternFill("solid", fgColor="DDDDDD")

        # 테두리 스타일
        title_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 병합된 영역 전체에 배경 + 테두리 적용
        for row in ws_dash["B30:P30"]:
            for cell in row:
                cell.fill = title_fill
                cell.border = title_border



        sending_chart = BarChart()
        sending_chart.title = None
        sending_chart.legend = None
        sending_chart.height = 12
        sending_chart.width = 28
        sending_chart.dataLabels = DataLabelList()
        sending_chart.dataLabels.showVal = True
        sending_chart.dataLabels.showCatName = False
        sending_chart.dataLabels.showSerName = False
        sending_chart.dataLabels.dLblPos = "outEnd"       

        sending_chart.add_data(
            Reference(
                ws_sending,
                min_col=2,
                min_row=ws_sending.max_row,         # 총합계 행
                max_col=ws_sending.max_column - 1,  # '총 건수' 제외
                max_row=ws_sending.max_row
            ),
            titles_from_data=False,
            from_rows=True  # ✅ 핵심
        )

        sending_chart.set_categories(
            Reference(
                ws_sending,
                min_col=2,
                min_row=1,
                max_col=ws_sending.max_column - 1,
                max_row=1
            )
        )

        # ✅ 막대 위 값 표시
        sending_chart.x_axis.tickLblPos = "nextTo"
        sending_chart.x_axis.majorTickMark = "out"
        sending_chart.x_axis.minorTickMark = "none"
        sending_chart.x_axis.delete = False

        ws_dash.add_chart(sending_chart, "B33")

    # -------------------------
    # 다운로드 버튼
    # -------------------------
    st.download_button(
        label="📥 누적 집계 엑셀 다운로드",
        data=output.getvalue(),
        file_name="픽업_샌딩_월별_시간대_누적집계.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("엑셀 파일을 하나 이상 업로드하세요.")
